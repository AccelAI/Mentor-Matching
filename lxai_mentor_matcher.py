#!/usr/bin/env python3
import csv
import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import time
from mentee import *
from mentor import *
from utils import *


if __name__ == '__main__':
    wb = Workbook()
    #create new xlsxwriter workbook object

    mentees = load_workbook('./Spring 2022/2022 Spring Cohort Mentees.xlsx')
    mentors = load_workbook('./Spring 2022/2022 Spring Cohort Mentors.xlsx')

    matches = wb.active

    matches.title = "Mentor to Mentee"

    menteeSheet = mentees.active
    mentorSheet = mentors.active

    row = 2
    allMentees = []
    allMentors = []

    while menteeSheet.cell(row=row, column=1).value != None and menteeSheet.cell(row=row, column=1).value != '':
        menteeRow = processMentee(menteeSheet, row)
        mentee = Mentee()
        for k,v in menteeRow.items():
            setattr(mentee, k, v)
    
        allMentees.append(mentee)
        row = row + 1
        # # If you want to see each mentee obj
        # mentee.printAll()

    row = 2
    while mentorSheet.cell(row=row, column=1).value != None and mentorSheet.cell(row=row, column=1).value != '':
        mentorRow = processMentor(mentorSheet, row)
        mentor = Mentor()
        for k,v in mentorRow.items():
            setattr(mentor, k, v)
    
        allMentors.append(mentor)
        row = row + 1
        # # If you want to see each mentor obj
        # mentor.printAll()

    #Statistics
    #Average Match
    menteeMatches = {}
    menteeAcceptableMentors = {}
    for mentee in allMentees:
        menteeId = mentee.menteeId
        menteePotentialMatches = mentorMatch(mentee, allMentors)        
        menteeMatches[menteeId] = menteePotentialMatches
        print("Working on mentee: {}".format(menteeId))
        menteeAcceptableMentors[menteeId] = acceptableMatches(menteePotentialMatches)


    priorityWithRating = menteePriority(allMentees)
    priority = []
    for item in priorityWithRating:
        priority.append(list(item.keys())[0])

    # Assign Mentees to Mentors based on mentee match % and mentors mentee limit
    unmatched = assignToMentor(menteeAcceptableMentors, priority, allMentors, allMentees)    

    if unmatched != {}:
        print("Not all mentees matched")
    else:
        print("All mentees matched")

    row = 1
    matches.cell(row=row, column=1).value = "ID #"
    matches.cell(row=row, column=1).font = Font(bold=True, size=12)
    matches.cell(row=row, column=1).alignment = Alignment(horizontal="left")    
    matches.cell(row=row, column=2).value = "Mentors"
    matches.cell(row=row, column=2).font = Font(bold=True, size=12)
    matches.cell(row=row, column=2).alignment = Alignment(horizontal="left")
    matches.cell(row=row, column=3).value = "Mentees"
    matches.cell(row=row, column=3).font = Font(bold=True, size=12)
    matches.cell(row=row, column=3).alignment = Alignment(horizontal="left")
    matches.cell(row=row, column=4).value = "Email"
    matches.cell(row=row, column=4).font = Font(bold=True, size=12)
    matches.cell(row=row, column=4).alignment = Alignment(horizontal="center")
    matches.cell(row=row, column=5).value = "Is LatinX"
    matches.cell(row=row, column=5).font = Font(bold=True, size=12)
    matches.cell(row=row, column=5).alignment = Alignment(horizontal="left")
    matches.cell(row=row, column=6).value = "Affiliation"
    matches.cell(row=row, column=6).font = Font(bold=True, size=12)
    matches.cell(row=row, column=6).alignment = Alignment(horizontal="left")
    matches.cell(row=row, column=7).value = "Seniority/Position"
    matches.cell(row=row, column=7).font = Font(bold=True, size=12)
    matches.cell(row=row, column=7).alignment = Alignment(horizontal="left")
    matches.cell(row=row, column=8).value = "Match"
    matches.cell(row=row, column=8).font = Font(bold=True, size=12)
    matches.cell(row=row, column=8).alignment = Alignment(horizontal="right")
    matches.cell(row=row, column=9).value = "Match Percents"
    matches.cell(row=row, column=9).font = Font(bold=True, size=12)
    matches.cell(row=row, column=9).alignment = Alignment(horizontal="right")
    matches.cell(row=row, column=10).value = "Conference Preference"
    matches.cell(row=row, column=10).font = Font(bold=True, size=12)
    matches.cell(row=row, column=10).alignment = Alignment(horizontal="left")
    matches.cell(row=row, column=11).value = "Mentoring Verticle"
    matches.cell(row=row, column=11).font = Font(bold=True, size=12)
    matches.cell(row=row, column=11).alignment = Alignment(horizontal="left")
    matches.cell(row=row, column=12).value = "Mentoring Skills"
    matches.cell(row=row, column=12).font = Font(bold=True, size=12)
    matches.cell(row=row, column=12).alignment = Alignment(horizontal="left")
    matches.cell(row=row, column=13).value = "Research Areas"
    matches.cell(row=row, column=13).font = Font(bold=True, size=12)
    matches.cell(row=row, column=13).alignment = Alignment(horizontal="left")
    matches.cell(row=row, column=14).value = "Languages"
    matches.cell(row=row, column=14).font = Font(bold=True, size=12)
    matches.cell(row=row, column=14).alignment = Alignment(horizontal="left")
    matches.cell(row=row, column=15).value = "Preferred Timzone"
    matches.cell(row=row, column=15).font = Font(bold=True, size=12)
    matches.cell(row=row, column=15).alignment = Alignment(horizontal="left")
    matches.cell(row=row, column=16).value = "Link to Website or Google Scholar"
    matches.cell(row=row, column=16).font = Font(bold=True, size=12)
    matches.cell(row=row, column=16).alignment = Alignment(horizontal="left")
    row = 2
    for mentor in allMentors:
        matches.cell(row=row, column=1).value = mentor.mentorId
        matches.cell(row=row, column=1).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=2).value = mentor.firstName + " " + mentor.lastName
        matches.cell(row=row, column=2).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=3).value = mentor.menteeLimit
        matches.cell(row=row, column=3).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=4).value = mentor.email
        matches.cell(row=row, column=4).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=5).value = mentor.isLatinx
        matches.cell(row=row, column=5).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=6).value = mentor.affiliation
        matches.cell(row=row, column=6).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=7).value = mentor.seniority
        matches.cell(row=row, column=7).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=8).value = ""
        matches.cell(row=row, column=8).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=9).value = ""
        matches.cell(row=row, column=9).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=10).value = str(mentor.mentorConfPref)
        matches.cell(row=row, column=10).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=11).value = str(mentor.mentoringVertical)
        matches.cell(row=row, column=11).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=12).value = str(mentor.mentoringSkills)
        matches.cell(row=row, column=12).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=13).value = str(mentor.researchAreas)
        matches.cell(row=row, column=13).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=14).value = str(mentor.languages)
        matches.cell(row=row, column=14).fill = PatternFill("solid", fgColor="DDDDDD")
        matches.cell(row=row, column=15).value = str(mentor.timezone)
        matches.cell(row=row, column=15).fill = PatternFill("solid", fgColor="DDDDDD")
        row += 1
        if len(mentor.mentorMatches) > 0:
            for mentee in mentor.mentorMatches:
                if mentee is not None:
                    matches.cell(row=row, column=1).value = mentee.menteeId
                    matches.cell(row=row, column=3).value = mentee.firstName + " " + mentee.lastName
                    matches.cell(row=row, column=4).value = mentee.email
                    matches.cell(row=row, column=5).value = mentee.isLatinx
                    matches.cell(row=row, column=6).value = mentee.affiliation
                    matches.cell(row=row, column=7).value = mentee.position
                    matches.cell(row=row, column=8).value = mentee.matchPercent
                    matches.cell(row=row, column=9).value = str(menteeAcceptableMentors[mentee.menteeId])
                    matches.cell(row=row, column=10).value = str(mentee.menteeConfPref)
                    matches.cell(row=row, column=11).value = str(mentee.mentoringVertical)
                    matches.cell(row=row, column=12).value = str(mentee.mentoringSkills)
                    matches.cell(row=row, column=13).value = str(mentee.researchAreas)
                    # matches.cell(row=row, column=14).value = str(mentee.careerAreas)
                    matches.cell(row=row, column=14).value = str(mentee.languages)
                    matches.cell(row=row, column=15).value = str(mentee.timezone)
                    matches.cell(row=row, column=16).value = str(mentee.website)
                    row += 1

    row += 1
    matches.cell(row=row, column=1).value = "Not Matched in last run"
    matches.cell(row=row, column=1).fill = PatternFill("solid", fgColor="DDDDDD")
    matches.cell(row=row, column=2).value = ""
    matches.cell(row=row, column=2).fill = PatternFill("solid", fgColor="DDDDDD")
    matches.cell(row=row, column=3).value = ""
    matches.cell(row=row, column=3).fill = PatternFill("solid", fgColor="DDDDDD")
    matches.cell(row=row, column=4).value = ""
    matches.cell(row=row, column=4).fill = PatternFill("solid", fgColor="DDDDDD")
    matches.cell(row=row, column=5).value = ""
    matches.cell(row=row, column=5).fill = PatternFill("solid", fgColor="DDDDDD")
    matches.cell(row=row, column=6).value = ""
    matches.cell(row=row, column=6).fill = PatternFill("solid", fgColor="DDDDDD")
    matches.cell(row=row, column=7).value = ""
    matches.cell(row=row, column=7).fill = PatternFill("solid", fgColor="DDDDDD")
    matches.cell(row=row, column=8).value = ""
    matches.cell(row=row, column=8).fill = PatternFill("solid", fgColor="DDDDDD")
    matches.cell(row=row, column=9).value = "Possible Matches (% Match:Mentor ID)"
    matches.cell(row=row, column=9).fill = PatternFill("solid", fgColor="DDDDDD")
    matches.cell(row=row, column=10).value = "Conference Preference"
    matches.cell(row=row, column=10).fill = PatternFill("solid", fgColor="DDDDDD")
    matches.cell(row=row, column=11).value = "Seeking Mentorship In"
    matches.cell(row=row, column=11).fill = PatternFill("solid", fgColor="DDDDDD")
    matches.cell(row=row, column=12).value = "Skills to be Mentored"
    matches.cell(row=row, column=12).fill = PatternFill("solid", fgColor="DDDDDD")
    matches.cell(row=row, column=13).value = "Research to be Mentored"
    matches.cell(row=row, column=13).fill = PatternFill("solid", fgColor="DDDDDD")
    row+=1

    if unmatched != {}:
        id_nums = sorted(unmatched.keys())
        for menteeid in id_nums:
            menteeid -= 2 # This is because allMentees is a list and mentee Ids start at 2, so two positions off
            matches.cell(row=row, column=1).value = menteeid
            matches.cell(row=row, column=3).value = allMentees[menteeid].firstName + " " + allMentees[menteeid].lastName
            matches.cell(row=row, column=4).value = allMentees[menteeid].email
            matches.cell(row=row, column=5).value = allMentees[menteeid].isLatinx
            matches.cell(row=row, column=6).value = allMentees[menteeid].affiliation
            matches.cell(row=row, column=7).value = allMentees[menteeid].position
            matches.cell(row=row, column=8).value = ""
            matches.cell(row=row, column=9).value = str(menteeAcceptableMentors[menteeid])
            matches.cell(row=row, column=10).value = str(allMentees[menteeid].menteeConfPref)
            matches.cell(row=row, column=11).value = str(allMentees[menteeid].mentoringVertical)
            matches.cell(row=row, column=12).value = str(allMentees[menteeid].mentoringSkills)
            matches.cell(row=row, column=13).value = str(allMentees[menteeid].researchAreas)
            # matches.cell(row=row, column=13).value = str(allMentees[menteeid].careerAreas)
            matches.cell(row=row, column=14).value = str(allMentees[menteeid].languages)
            matches.cell(row=row, column=15).value = str(allMentees[menteeid].timezone)
            matches.cell(row=row, column=16).value = str(allMentees[menteeid].website)
            row += 1

    wb.save("./Spring 2022/2022 Spring Cohort Matched.xlsx")
